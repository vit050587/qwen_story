import os
import re
import json
import base64
import time
import fitz
import pandas as pd
from pathlib import Path
from datetime import datetime
from ollama import Client
from openpyxl.styles import Font, Alignment, PatternFill
from .config import load_config

def get_ollama_client():
    config = load_config()
    return Client(host=config.ollama_url)

def pdf_to_base64(pdf_path: str) -> str:
    """Конвертирует первую страницу PDF в base64 PNG"""
    try:
        doc = fitz.open(pdf_path)
        page = doc[0]
        # Zoom 2x для лучшего качества
        mat = fitz.Matrix(2, 2) 
        pix = page.get_pixmap(matrix=mat)
        data = pix.tobytes("png")
        doc.close()
        return base64.b64encode(data).decode("utf-8")
    except Exception as e:
        print(f"❌ Ошибка конвертации изображения {pdf_path}: {e}")
        return None

def create_analysis_prompt() -> str:
    return """Ты эксперт по анализу архитектурных чертежей. Ответь строго по пунктам (1-13) без лишнего текста, маркдауна и вступлений.
Если параметр неприменим или не виден, пиши '-'.

1. Масштаб (например: 1:100 или 'не указан').
2. Тип чертежа (вид сбоку всего здания / вид сверху этажа / другое).
3. Общая высота здания в метрах (если вид сбоку), иначе '-'.
4. Тип этажа (жилой / технический / парковка / подвал / чердак / не указан), если вид сверху, иначе '-'.
5. Номер этажа (первый / второй / типовой / технический), если вид сверху, иначе '-'.
6. Общая площадь квартир (м²), если этаж жилой и площадь указана, иначе '-'.
7. Количество эвакуационных выходов на улицу (только для 1 этажа), иначе '-'.
8. Количество лестничных клеток для эвакуации (если этаж жилой), иначе '-'.
9. Ширина лестничного марша в метрах (точность 0.05), если известен масштаб и этаж жилой, иначе '-'.
10. Номера квартир с балконами/лоджиями (если этаж жилой), перечисли через запятую, иначе '-'.
11. Расстояние от двери самой дальней квартиры до ближайшей лестничной клетки (м), если известно, иначе '-'.
12. Общая длина основного коридора (м), если известно, иначе '-'.
13. Ширина коридора в самом узком месте (м), если известно, иначе '-'.

Пример правильного ответа:
1. 1:100
2. вид сверху этажа
3. -
4. жилой
5. типовой
6. 450
7. -
8. 2
9. 1.35
10. кв. 5, 8, 12
11. 15
12. 24
13. 1.5
"""

def analyze_page_vlm(page_path: str, client: Client) -> dict:
    """Отправляет чертеж модели и возвращает распарсенный ответ"""
    config = load_config()
    b64 = pdf_to_base64(page_path)
    if not b64:
        return None

    prompt = create_analysis_prompt()
    
    try:
        resp = client.chat(
            model=config.drawing_vlm_model,
            messages=[{'role': 'user', 'content': prompt, 'images': [b64]}],
            options={'temperature': 0.0, 'num_predict': 2000}
        )
        
        text = ""
        if hasattr(resp, 'message'):
            text = resp.message.content
        elif isinstance(resp, dict) and 'message' in resp:
            text = resp['message']['content']
            
        # Парсинг ответа
        result = {}
        for line in text.split('\n'):
            m = re.match(r'(\d+)\.\s*(.+)', line.strip())
            if m:
                result[int(m.group(1))] = m.group(2).strip()
        
        return result if result else None
        
    except Exception as e:
        print(f"❌ Ошибка API при анализе страницы {page_path}: {e}")
        return None

def validate_data_with_llm(data: dict, client: Client) -> dict:
    """Генерирует проверку соответствия нормам через LLM"""
    config = load_config()
    
    # Формируем краткую выжимку данных
    data_summary = "\n".join([f"{k}: {v}" for k, v in data.items()])
    
    prompt = f"""Ты эксперт по пожарной безопасности (СП 1.13130). 
Проверь извлеченные данные чертежа на соответствие нормам.
Данные:
{data_summary}

Верни ТОЛЬКО JSON объект со следующими ключами и структурой:
{{
  "balconies": {{"status": "ok/not_ok/not_applicable", "value": "значение", "error_text": "пояснение"}},
  "staircases": {{"status": "...", "value": "...", "error_text": "..."}},
  "distance_to_exit": {{"status": "...", "value": "...", "error_text": "..."}},
  "corridor_width": {{"status": "...", "value": "...", "error_text": "..."}},
  "stair_width": {{"status": "...", "value": "...", "error_text": "..."}},
  "building_height": {{"status": "...", "value": "...", "error_text": "..."}},
  "floor_type": {{"status": "...", "value": "...", "error_text": "..."}}
}}

Логика проверки:
- balconies: наличие балконов в секционных домах.
- staircases: кол-во лестничных клеток (минимум 2 для этажа > 500м2 или > 12 квартир).
- distance_to_exit: расстояние до выхода не более нормируемого.
- corridor_width: ширина коридора не менее 1.4м (для жилых).
- stair_width: ширина марша не менее 1.05м.
- building_height: проверка классификации по высоте.
- floor_type: корректность типа этажа.

Если данных недостаточно для проверки, ставь status: "not_applicable".
"""

    try:
        resp = client.chat(
            model=config.drawing_validation_model,
            messages=[{'role': 'user', 'content': prompt}],
            options={'temperature': 0.0, 'num_predict': 1500}
        )
        
        text = ""
        if hasattr(resp, 'message'):
            text = resp.message.content
        elif isinstance(resp, dict) and 'message' in resp:
            text = resp['message']['content']
            
        # Извлекаем JSON из ответа
        match = re.search(r'\{[\s\S]*\}', text)
        if match:
            return json.loads(match.group())
        return None
        
    except Exception as e:
        print(f"⚠️ Ошибка валидации: {e}")
        # Возвращаем заглушку, чтобы не ломать процесс
        return {
            'balconies': {'status': 'not_applicable', 'value': data.get(10, '-'), 'error_text': 'Ошибка валидации'},
            'staircases': {'status': 'not_applicable', 'value': data.get(8, '-'), 'error_text': ''},
            'distance_to_exit': {'status': 'not_applicable', 'value': data.get(11, '-'), 'error_text': ''},
            'corridor_width': {'status': 'not_applicable', 'value': data.get(13, '-'), 'error_text': ''},
            'stair_width': {'status': 'not_applicable', 'value': data.get(9, '-'), 'error_text': ''},
            'building_height': {'status': 'not_applicable', 'value': data.get(3, '-'), 'error_text': ''},
            'floor_type': {'status': 'not_applicable', 'value': data.get(4, '-'), 'error_text': ''}
        }

def create_excel_report(results: list, out_path: str):
    """Создает итоговый Excel файл"""
    checks_map = [
        ('balconies', 'Наличие балконов/лоджий'), 
        ('staircases', 'Количество лестничных клеток'),
        ('distance_to_exit', 'Расстояние до эвакуационного выхода'), 
        ('corridor_width', 'Ширина коридора'),
        ('stair_width', 'Ширина лестничного марша'), 
        ('building_height', 'Высота здания'),
        ('floor_type', 'Тип этажа')
    ]
    
    rows = []
    for key, name in checks_map:
        val = "-"
        status = "not_found"
        comment = ""
        src_pages = []
        
        # Агрегируем данные со всех страниц
        for r in results:
            v_data = r.get('validation', {}).get(key, {})
            v_val = v_data.get('value', '-')
            
            # Берем первое непустое значение
            if v_val and v_val != '-' and val == "-":
                val = v_val
                status = v_data.get('status', 'not_found')
                comment = v_data.get('error_text', '')
            
            # Собираем все страницы, где найдено значение
            if v_val and v_val != '-':
                src_pages.append(str(r['page']))
        
        rows.append({
            'Параметр': name,
            'Значение': val,
            'Статус': status,
            'Комментарий': comment,
            'Источник (страницы)': ", ".join(src_pages) if src_pages else "-"
        })
    
    df = pd.DataFrame(rows)
    
    # Создаем Excel с форматированием
    with pd.ExcelWriter(out_path, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Анализ чертежей')
        ws = writer.sheets['Анализ чертежей']
        
        # Настройка ширины колонок
        col_widths = {'A': 30, 'B': 15, 'C': 15, 'D': 40, 'E': 20}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
            
        # Стилизация заголовков
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            
        # Цветовая индикация статусов
        for row_idx in range(2, len(df) + 2):
            status_cell = ws.cell(row=row_idx, column=3) # Колонка C (Статус)
            status_val = status_cell.value
            
            if status_val == 'ok':
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
            elif status_val == 'not_ok':
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006", bold=True)
            elif status_val == 'not_applicable':
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                
    print(f"📊 Excel отчет сохранен: {out_path}")

def run_analysis(session_id: str, detected_drawings: list, output_dir: str) -> str:
    """
    ТОЧКА ВХОДА для этапа анализа.
    Принимает список словарей от детектора.
    Возвращает путь к созданному Excel файлу или None.
    """
    if not detected_drawings:
        return None
        
    config = load_config()
    client = get_ollama_client()
    print(f"🧠 Запуск анализа {len(detected_drawings)} чертежей моделью {config.drawing_vlm_model}...")
    
    results = []
    
    for i, dw in enumerate(detected_drawings):
        print(f"   Обработка страницы {dw['page_num']} ({i+1}/{len(detected_drawings)})...")
        
        parsed_data = analyze_page_vlm(dw['file_path'], client)
        
        if parsed_data:
            print(f"      Данные извлечены. Валидация...")
            validation = validate_data_with_llm(parsed_data, client)
            
            results.append({
                'page': dw['page_num'],
                'parsed': parsed_data,
                'validation': validation
            })
        else:
            print(f"      ⚠️ Не удалось извлечь данные")
        
        # Пауза между запросами
        if i < len(detected_drawings) - 1:
            time.sleep(2)
    
    if not results:
        print("⚠️ Ни один чертеж не был успешно проанализирован.")
        return None
    
    # Генерация отчета
    report_filename = f"drawings_report_{session_id}.xlsx"
    report_path = Path(output_dir) / report_filename
    create_excel_report(results, str(report_path))
    
    return str(report_path)